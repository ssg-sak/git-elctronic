"""Build Excel dashboard with charts — personal portfolio only (not handoff).

Outputs under 포폴용_개인_대시보드_20260809/ (+ Desktop mirror).
"""
from __future__ import annotations

import json
import shutil
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

REPO = Path(__file__).resolve().parents[5]
PORT = REPO / "포폴용_개인_대시보드_20260809"
DATA = PORT / "data"
OUT_XLSX = PORT / "DA1_대시보드_20260809.xlsx"
DESK_DIR = Path.home() / "Desktop" / PORT.name
DESK_XLSX = DESK_DIR / "DA1_대시보드_20260809.xlsx"

HEADER = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True, name="Malgun Gothic", size=11)
TITLE = Font(name="Malgun Gothic", size=16, bold=True, color="1F4E79")
SUB = Font(name="Malgun Gothic", size=10, color="666666")
THIN = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)


def _read(name: str) -> pd.DataFrame:
    p = DATA / name
    if not p.is_file():
        return pd.DataFrame()
    return pd.read_csv(p, encoding="utf-8-sig")


def _style_header(ws, row=1, cols=1):
    for c in range(1, cols + 1):
        cell = ws.cell(row, c)
        cell.fill = HEADER
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN


def _write_df(ws, df: pd.DataFrame, start_row=1, start_col=1):
    if df.empty:
        ws.cell(start_row, start_col, "(데이터 없음)")
        return 0, 0
    for r_i, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
        for c_i, val in enumerate(row, start=start_col):
            cell = ws.cell(r_i, c_i, val)
            cell.border = THIN
            cell.font = Font(name="Malgun Gothic", size=9)
    _style_header(ws, start_row, len(df.columns))
    for i, col in enumerate(df.columns, start=start_col):
        ws.column_dimensions[get_column_letter(i)].width = min(22, max(10, len(str(col)) + 4))
    return len(df) + 1, len(df.columns)


def main() -> None:
    meta = _read("dim_meta.csv")
    hour = _read("fact_eda_hour.csv")
    dow = _read("fact_eda_dow.csv")
    fresh = _read("fact_eda_freshness.csv")
    charger = _read("fact_eda_charger_bucket.csv")
    kpi = _read("fact_kpi.csv")
    loo = _read("fact_loo_deltas.csv")
    feats = _read("dim_final_features.csv")
    over = _read("fact_overfit_summary.csv")

    as_of = str(meta["as_of"].iloc[0]) if len(meta) and "as_of" in meta.columns else "?"
    pull = str(meta["pull"].iloc[0]) if len(meta) and "pull" in meta.columns else "?"

    wb = Workbook()

    # --- Overview ---
    ws0 = wb.active
    ws0.title = "00_한눈에"
    ws0["A1"] = "EV SafeCharge DA① 대시보드 (20260809)"
    ws0["A1"].font = TITLE
    ws0.merge_cells("A1:F1")
    ws0["A2"] = f"현재표 as_of: {as_of}  |  pull: {pull}  |  점수 없음(②)"
    ws0["A2"].font = SUB
    ws0["A4"] = "이 파일이 대시보드입니다. 아래 시트 탭만 보면 됩니다."
    ws0["A4"].font = Font(name="Malgun Gothic", size=12, bold=True)
    ws0["A6"] = "시트 안내"
    ws0["A6"].font = Font(name="Malgun Gothic", size=12, bold=True, color="1F4E79")
    guides = [
        ("01_시간대가용", "몇 시에 비나 (E1)"),
        ("02_요일가용", "요일별 가용 · 잠정 (E2)"),
        ("03_신선도", "HIGH/NORMAL/CHECK 소 수 (E4)"),
        ("04_충전기대수", "대수 버킷별 가용 (E3)"),
        ("05_피처LOO", "피처 빼면 성능 얼마나 깨지나"),
        ("06_KPI", "운영 KPI 표"),
        ("07_최종피처", "최종 9개"),
        ("08_과적합요약", "과적합 요약"),
    ]
    ws0["A7"] = "시트"
    ws0["B7"] = "내용"
    _style_header(ws0, 7, 2)
    for i, (a, b) in enumerate(guides, start=8):
        ws0.cell(i, 1, a).font = Font(name="Malgun Gothic", size=10)
        ws0.cell(i, 2, b).font = Font(name="Malgun Gothic", size=10)
    ws0.column_dimensions["A"].width = 18
    ws0.column_dimensions["B"].width = 40
    ws0["A18"] = "포폴 전용(전달 아님). Power BI: 데이터 가져오기 → 이 폴더\\data"
    ws0["A18"].font = SUB

    def sheet_chart(title: str, df: pd.DataFrame, xcol: str, ycol: str, chart_type: str):
        ws = wb.create_sheet(title)
        if df.empty or xcol not in df.columns or ycol not in df.columns:
            ws["A1"] = f"{title}: 데이터 없음"
            return
        use = df[[xcol, ycol]].copy()
        use[ycol] = pd.to_numeric(use[ycol], errors="coerce")
        use = use.dropna()
        nrows, ncols = _write_df(ws, use, 1, 1)
        if chart_type == "line":
            chart = LineChart()
        else:
            chart = BarChart()
        chart.title = title
        chart.style = 10
        chart.y_axis.title = ycol
        chart.x_axis.title = xcol
        data_ref = Reference(ws, min_col=2, min_row=1, max_row=nrows, max_col=2)
        cats = Reference(ws, min_col=1, min_row=2, max_row=nrows)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats)
        chart.shape = 4
        chart.width = 18
        chart.height = 10
        ws.add_chart(chart, "D2")

    # normalize column names for hour
    if not hour.empty:
        if "hour" not in hour.columns and "\ufeffhour" in hour.columns:
            hour = hour.rename(columns={"\ufeffhour": "hour"})
        sheet_chart("01_시간대가용", hour, "hour", "avail_mean", "line")
    else:
        wb.create_sheet("01_시간대가용")["A1"] = "없음"

    if not dow.empty:
        x = "dow_ko" if "dow_ko" in dow.columns else "dow"
        sheet_chart("02_요일가용", dow, x, "avail_mean", "bar")
    else:
        wb.create_sheet("02_요일가용")["A1"] = "없음"

    if not fresh.empty:
        x = "age_bucket" if "age_bucket" in fresh.columns else fresh.columns[0]
        y = "stations" if "stations" in fresh.columns else fresh.columns[1]
        sheet_chart("03_신선도", fresh, x, y, "bar")
    else:
        wb.create_sheet("03_신선도")["A1"] = "없음"

    if not charger.empty:
        x = "bucket" if "bucket" in charger.columns else charger.columns[0]
        y = "d1_avail_mean_observed" if "d1_avail_mean_observed" in charger.columns else "d2_avail_mean"
        if y in charger.columns:
            sheet_chart("04_충전기대수", charger, x, y, "bar")
        else:
            wb.create_sheet("04_충전기대수")["A1"] = "컬럼 없음"
    else:
        wb.create_sheet("04_충전기대수")["A1"] = "없음"

    # LOO
    ws = wb.create_sheet("05_피처LOO")
    if not loo.empty:
        feat_col = next((c for c in ("feature", "removed", "col", "name") if c in loo.columns), loo.columns[0])
        metric_col = next(
            (c for c in ("delta_pr_auc", "pr_auc_delta", "delta_prauc") if c in loo.columns),
            None,
        )
        if metric_col is None:
            nums = [c for c in loo.columns if pd.api.types.is_numeric_dtype(loo[c])]
            metric_col = nums[0] if nums else None
        if metric_col:
            use = loo[[feat_col, metric_col]].dropna().sort_values(metric_col)
            nrows, _ = _write_df(ws, use, 1, 1)
            chart = BarChart()
            chart.title = "피처 제거 Δ"
            chart.style = 10
            data_ref = Reference(ws, min_col=2, min_row=1, max_row=nrows, max_col=2)
            cats = Reference(ws, min_col=1, min_row=2, max_row=nrows)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats)
            chart.width = 18
            chart.height = 10
            ws.add_chart(chart, "D2")
        else:
            _write_df(ws, loo, 1, 1)
    else:
        ws["A1"] = "LOO 없음"

    ws = wb.create_sheet("06_KPI")
    _write_df(ws, kpi, 1, 1)

    ws = wb.create_sheet("07_최종피처")
    _write_df(ws, feats, 1, 1)

    ws = wb.create_sheet("08_과적합요약")
    _write_df(ws, over, 1, 1)

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)
    DESK_DIR.mkdir(parents=True, exist_ok=True)
    shutil.copy2(OUT_XLSX, DESK_XLSX)

    # open
    import os

    os.startfile(str(OUT_XLSX))  # noqa: S606
    html = PORT / "대시보드_바로보기.html"
    if html.is_file():
        os.startfile(str(html))  # noqa: S606

    print(
        json.dumps(
            {
                "ok": True,
                "xlsx": str(OUT_XLSX),
                "desktop": str(DESK_XLSX),
                "handoff": False,
            },
            ensure_ascii=True,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
